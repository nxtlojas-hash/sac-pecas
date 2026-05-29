"""Extrai leads de J:/Meu Drive/Leads_RespondIO.xlsx para JSON."""
from __future__ import annotations
import json
from pathlib import Path
from openpyxl import load_workbook

SOURCE = "J:/Meu Drive/Leads_RespondIO.xlsx"
OUT = Path(__file__).parent / "data" / "leads_respondio.json"


def main() -> None:
    wb = load_workbook(SOURCE, read_only=True, data_only=True)
    ws = wb["Leads - Respond.io"]
    leads = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        leads.append({
            "nome": row[0] or "",
            "sobrenome": row[1] or "",
            "telefone": row[2] or "",
            "modelo": row[3] or "",
            "cidade": row[4] or "",
            "estado": row[5] or "",
            "lifecycle": row[6] or "",
            "status": row[7] or "",
        })
    wb.close()
    OUT.write_text(json.dumps(leads, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(f"Extraidos {len(leads)} leads -> {OUT}")


if __name__ == "__main__":
    main()
