"""Generate the Controle - Motos Sumaré.xlsx workbook with 4 sheets."""
from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from tools.sumare_import.schema import Moto, MOTOS_NXT, STATUS_OPCOES, TIPO_ATENDIMENTO

MOTOS_HEADERS = [
    "ID", "Data registro", "Nº OS", "Data entrada checklist", "Status atual",
    "Identificada?", "Nome", "CPF", "Telefone", "Endereço", "Cidade/UF",
    "Modelo NXT", "Cor", "Nº Chassi", "Nº Motor", "Data compra", "Loja/revendedor",
    "Problema relatado", "Componentes danificados", "Tipo atendimento",
    "Peças substituídas", "Técnico responsável",
    "Foto checklist", "Foto moto", "Fotos extras",
    "Origem do registro", "Quem registrou", "Observações",
    # Cruzamento WhatsApp 29/05
    "Categoria", "WA 29/05 match", "WA Motivo 29/05", "WA Cidade 29/05",
    "WA Conflitos",
    # Cruzamento CASOS SAC E SUMARE
    "SAC Aguardando match", "SAC Status", "SAC Peças solicitadas",
    "SAC Peças match", "SAC Resumo peças",
]
assert len(MOTOS_HEADERS) == 38

MOV_HEADERS = ["Data/hora", "ID Moto", "Status anterior", "Status novo", "Responsável", "Observação"]
DASH_HEADERS = ["Métrica", "Valor"]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_headers(ws, headers: list[str]) -> None:
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
    ws.freeze_panes = "A2"


def _moto_row(m: Moto, extras: dict | None = None) -> list:
    extras = extras or {}
    return [
        m.id, m.data_registro, m.numero_os, m.data_entrada_checklist, m.status_atual,
        "Sim" if m.identificada() else "Não",
        m.nome, m.cpf, m.telefone, m.endereco, m.cidade_uf,
        m.modelo_nxt, m.cor, m.chassi, m.motor, m.data_compra, m.loja,
        m.problema_relatado, "; ".join(m.componentes_danificados),
        m.tipo_atendimento, m.pecas_substituidas, m.tecnico_responsavel,
        m.foto_checklist, m.foto_moto, ", ".join(m.fotos_extras),
        m.origem_registro, m.quem_registrou, m.observacoes,
        # Extras (cruzamento)
        extras.get("categoria", ""),
        extras.get("wa_match", ""),
        extras.get("wa_motivo_29_05", "") or "",
        extras.get("wa_cidade_29_05", "") or "",
        extras.get("wa_conflitos", "") or "",
        extras.get("sac_match_aguardando", ""),
        extras.get("sac_status_aguardando", "") or "",
        extras.get("sac_pecas_solicitadas", "") or "",
        extras.get("sac_match_pecas", ""),
        extras.get("sac_resumo_pecas", "") or "",
    ]


def _add_validations(ws, motos_count: int) -> None:
    end = max(motos_count + 1, 200)
    def rng(col: int) -> str:
        return f"{get_column_letter(col)}2:{get_column_letter(col)}{end}"

    dv_status = DataValidation(type="list", formula1='"' + ",".join(STATUS_OPCOES) + '"', allow_blank=True)
    dv_status.add(rng(5))
    ws.add_data_validation(dv_status)

    dv_id = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=True)
    dv_id.add(rng(6))
    ws.add_data_validation(dv_id)

    dv_modelo = DataValidation(type="list", formula1='"' + ",".join(MOTOS_NXT) + '"', allow_blank=True)
    dv_modelo.add(rng(12))
    ws.add_data_validation(dv_modelo)

    dv_tipo = DataValidation(type="list", formula1='"' + ",".join(TIPO_ATENDIMENTO) + '"', allow_blank=True)
    dv_tipo.add(rng(20))
    ws.add_data_validation(dv_tipo)


def build_workbook(
    motos: list[Moto],
    path: Path,
    extras_by_id: dict[str, dict] | None = None,
    cemiterio: list[dict] | None = None,
    novas_wa: list[dict] | None = None,
) -> None:
    extras_by_id = extras_by_id or {}
    wb = Workbook()
    # --- Aba Motos ---
    ws_motos = wb.active
    ws_motos.title = "Motos"
    _write_headers(ws_motos, MOTOS_HEADERS)
    for m in motos:
        ws_motos.append(_moto_row(m, extras_by_id.get(m.id)))
    _add_validations(ws_motos, len(motos))
    widths = [8, 12, 10, 14, 22, 13, 28, 18, 16, 38, 16, 14, 14, 16, 16, 12, 22, 38, 32, 18, 28, 22, 32, 32, 22, 22, 22, 38,
              22, 13, 38, 16, 32, 16, 18, 32, 16, 32]
    for i, w in enumerate(widths, start=1):
        ws_motos.column_dimensions[get_column_letter(i)].width = w

    # --- Aba Movimentações ---
    ws_mov = wb.create_sheet("Movimentações")
    _write_headers(ws_mov, MOV_HEADERS)
    for m in motos:
        ws_mov.append([
            f"{m.data_registro} 13:00",
            m.id,
            "",
            "Aguardando diagnóstico",
            m.quem_registrou or "",
            "Recebida — levantamento pós-incêndio 28/05",
        ])
    dv_mov = DataValidation(type="list", formula1='"' + ",".join(STATUS_OPCOES) + '"', allow_blank=True)
    end_mov = max(len(motos) + 1, 500)
    dv_mov.add(f"C2:D{end_mov}")
    ws_mov.add_data_validation(dv_mov)
    for i, w in enumerate([18, 10, 22, 22, 22, 50], start=1):
        ws_mov.column_dimensions[get_column_letter(i)].width = w

    # --- Aba Não Identificadas (dados estáticos, snapshot do build) ---
    ws_ni = wb.create_sheet("Não Identificadas")
    ws_ni["A1"] = "Snapshot das motos não identificadas no momento do build. Re-rodar tools/sumare_import/run.py para regenerar."
    ws_ni["A1"].font = Font(italic=True, color="808080")
    ni_headers = ["ID", "Nome (parcial)", "Modelo NXT", "Cor", "Problema relatado", "Foto checklist", "Foto moto", "Pistas (preencher conforme investiga)"]
    for col_idx, h in enumerate(ni_headers, start=1):
        c = ws_ni.cell(row=3, column=col_idx, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
    row = 4
    for m in motos:
        if m.identificada():
            continue
        ws_ni.cell(row=row, column=1, value=m.id)
        ws_ni.cell(row=row, column=2, value=m.nome or "")
        ws_ni.cell(row=row, column=3, value=m.modelo_nxt or "")
        ws_ni.cell(row=row, column=4, value=m.cor or "")
        ws_ni.cell(row=row, column=5, value=m.problema_relatado or "")
        ws_ni.cell(row=row, column=6, value=m.foto_checklist or "")
        ws_ni.cell(row=row, column=7, value=m.foto_moto or "")
        row += 1
    for i, w in enumerate([8, 26, 14, 14, 38, 32, 32, 32], start=1):
        ws_ni.column_dimensions[get_column_letter(i)].width = w
    ws_ni.freeze_panes = "A4"

    # --- Aba Dashboard ---
    ws_d = wb.create_sheet("Dashboard")
    _write_headers(ws_d, DASH_HEADERS)
    # Calcula valores no build pra garantir funcionar mesmo se as formulas falharem na conversao
    total = len(motos)
    identificadas = sum(1 for m in motos if m.identificada())
    nao_id = total - identificadas
    ws_d.append(["Total motos", total])
    ws_d.append(["Identificadas", identificadas])
    ws_d.append(["Não identificadas", nao_id])
    ws_d.append(["% identificadas", f"{identificadas*100//max(total,1)}%"])
    ws_d.append([])
    ws_d.append(["Status", "Quantidade"])
    for s in STATUS_OPCOES:
        count = sum(1 for m in motos if m.status_atual == s)
        ws_d.append([s, count])
    ws_d.append([])
    ws_d.append(["Modelo", "Quantidade"])
    for m_nome in MOTOS_NXT:
        count = sum(1 for m in motos if m.modelo_nxt == m_nome)
        ws_d.append([m_nome, count])
    ws_d.column_dimensions["A"].width = 28
    ws_d.column_dimensions["B"].width = 16

    # --- Aba Motos no Galpão (visão consolidada 29/05) ---
    ws_g = wb.create_sheet("Motos no Galpão")
    galpao_headers = [
        "Origem", "ID nosso", "Cliente", "Telefone (4 últ)", "Modelo", "Cor",
        "Cidade", "Motivo / Problema atual", "Categoria", "Foto checklist", "Foto moto"
    ]
    for col_idx, h in enumerate(galpao_headers, start=1):
        c = ws_g.cell(row=1, column=col_idx, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
    ws_g.freeze_panes = "A2"
    # Linhas: 1 por moto da nossa planilha (com info do WA quando matched)
    for m in motos:
        ex = extras_by_id.get(m.id, {})
        if ex.get("categoria") == "Cemitério / Estoque NXT":
            continue  # cemiterio vai para aba dedicada
        ws_g.append([
            "in-loco + WA" if ex.get("wa_match") == "Sim" else "in-loco apenas",
            m.id,
            m.nome or "",
            m.telefone[-4:] if m.telefone and len(m.telefone) >= 4 else "",
            m.modelo_nxt or "",
            m.cor or "",
            m.cidade_uf or "",
            ex.get("wa_motivo_29_05") or m.problema_relatado or "",
            ex.get("categoria") or "Cliente (assistência)",
            m.foto_checklist or "",
            m.foto_moto or "",
        ])
    # Linhas: motos do WA que NÃO bateram (novas no galpão pós-28/05)
    for wa in (novas_wa or []):
        ws_g.append([
            "Só WA 29/05",
            "",
            wa.get("cliente") or "",
            wa.get("cel_4dig") or "",
            wa.get("modelo") or "",
            wa.get("cor") or "",
            wa.get("cidade") or "",
            wa.get("motivo") or "",
            "Cliente (assistência) - novo",
            "",
            "",
        ])
    for i, w in enumerate([18, 10, 28, 14, 16, 14, 18, 50, 24, 32, 32], start=1):
        ws_g.column_dimensions[get_column_letter(i)].width = w

    # --- Aba Cemitério ---
    ws_c = wb.create_sheet("Cemitério")
    ws_c["A1"] = "Motos da NXT canibalizadas para retirar peças e atender outras assistências."
    ws_c["A1"].font = Font(italic=True, color="808080")
    cem_headers = ["ID", "Modelo", "Cor", "Motivo / Peças retiradas"]
    for col_idx, h in enumerate(cem_headers, start=1):
        c = ws_c.cell(row=3, column=col_idx, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
    for idx, entry in enumerate((cemiterio or []), start=1):
        ws_c.append([
            f"C{idx:03d}",
            entry.get("modelo") or "",
            entry.get("cor") or "",
            entry.get("motivo") or "",
        ])
    for i, w in enumerate([8, 18, 16, 52], start=1):
        ws_c.column_dimensions[get_column_letter(i)].width = w

    # --- Aba Estoque Total Sumaré 29/05 (snapshot consolidado) ---
    ws_e = wb.create_sheet("Estoque Total Sumaré 29-05")
    ws_e["A1"] = "Snapshot de TODAS as motos no galpão de Sumaré em 29/05/2026 (3 fontes consolidadas)."
    ws_e["A1"].font = Font(italic=True, color="808080")
    est_headers = [
        "ID", "Origem", "Categoria", "Cliente", "Telefone", "Cidade",
        "Modelo", "Cor", "Motivo / Problema atual",
        "Foto checklist", "Foto moto", "Status"
    ]
    for col_idx, h in enumerate(est_headers, start=1):
        c = ws_e.cell(row=3, column=col_idx, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN

    # 1. Todas as 75 motos in-loco
    for m in motos:
        ex = extras_by_id.get(m.id, {})
        is_cemiterio = ex.get("categoria") == "Cemitério / Estoque NXT"
        ws_e.append([
            m.id,
            "in-loco + WA" if ex.get("wa_match") == "Sim" else ("in-loco apenas" if not is_cemiterio else "in-loco (cemitério)"),
            ex.get("categoria") or "Cliente (assistência)",
            m.nome or "",
            m.telefone[-4:] if m.telefone and len(m.telefone) >= 4 else "",
            ex.get("wa_cidade_29_05") or m.cidade_uf or "",
            m.modelo_nxt or "",
            m.cor or "",
            ex.get("wa_motivo_29_05") or m.problema_relatado or "",
            m.foto_checklist or "",
            m.foto_moto or "",
            m.status_atual,
        ])

    # 2. WA novas (assistência sem match in-loco)
    for idx, wa in enumerate((novas_wa or []), start=1):
        ws_e.append([
            f"W{idx:03d}",
            "Só WA 29/05",
            "Cliente (assistência) - novo",
            wa.get("cliente") or "",
            wa.get("cel_4dig") or "",
            wa.get("cidade") or "",
            wa.get("modelo") or "",
            wa.get("cor") or "",
            wa.get("motivo") or "",
            "",
            "",
            "Aguardando diagnóstico",
        ])

    # 3. Cemitério (18 motos NXT pra retirada de peças)
    for idx, entry in enumerate((cemiterio or []), start=1):
        ws_e.append([
            f"C{idx:03d}",
            "Cemitério WA",
            "Cemitério / Estoque NXT",
            "NXT (cemitério)",
            "",
            "",
            entry.get("modelo") or "",
            entry.get("cor") or "",
            entry.get("motivo") or "",
            "",
            "",
            "Cemitério - retirada de peças",
        ])

    ws_e.freeze_panes = "A4"
    for i, w in enumerate([8, 18, 24, 28, 14, 18, 16, 14, 50, 32, 32, 22], start=1):
        ws_e.column_dimensions[get_column_letter(i)].width = w

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    import json
    data_dir = Path(__file__).parent / "data"
    motos_json = data_dir / "motos.json"
    raw = json.loads(motos_json.read_text(encoding="utf-8"))
    # filtrar campos que não são do dataclass Moto
    moto_fields = set(Moto.__dataclass_fields__.keys())
    motos = [Moto(**{k: v for k, v in d.items() if k in moto_fields}) for d in raw]
    out = Path(__file__).parent / "out" / "Controle - Motos Sumaré.xlsx"
    build_workbook(motos, out)
    print(f"Wrote {out} with {len(motos)} motos.")


if __name__ == "__main__":
    main()
