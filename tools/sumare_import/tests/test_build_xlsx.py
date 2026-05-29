from pathlib import Path
from openpyxl import load_workbook
from tools.sumare_import.schema import Moto
from tools.sumare_import.build_xlsx import build_workbook


def test_workbook_has_7_sheets(tmp_path: Path):
    out = tmp_path / "test.xlsx"
    build_workbook(motos=[], path=out)
    wb = load_workbook(out)
    assert set(wb.sheetnames) == {
        "Motos", "Movimentações", "Não Identificadas", "Dashboard",
        "Motos no Galpão", "Cemitério", "Estoque Total Sumaré 29-05",
    }


def test_motos_sheet_has_38_columns_header(tmp_path: Path):
    out = tmp_path / "test.xlsx"
    build_workbook(motos=[], path=out)
    wb = load_workbook(out)
    ws = wb["Motos"]
    headers = [c.value for c in ws[1]]
    assert headers[0] == "ID"
    assert "CPF" in headers
    assert "Foto checklist" in headers
    assert "WA 29/05 match" in headers
    assert "SAC Aguardando match" in headers
    assert "Categoria" in headers
    assert len([h for h in headers if h]) == 38


def test_motos_sheet_populates_data_rows(tmp_path: Path):
    out = tmp_path / "test.xlsx"
    moto = Moto(
        id="M001",
        data_registro="2026-05-28",
        nome="Nathalia do Rego Martins",
        cpf="099.194.986-00",
        modelo_nxt="Jaya",
    )
    build_workbook(motos=[moto], path=out)
    wb = load_workbook(out)
    ws = wb["Motos"]
    assert ws["A2"].value == "M001"
    headers = [c.value for c in ws[1]]
    nome_col_idx = headers.index("Nome") + 1
    assert ws.cell(row=2, column=nome_col_idx).value == "Nathalia do Rego Martins"


def test_movimentacoes_pre_populated(tmp_path: Path):
    out = tmp_path / "test.xlsx"
    moto = Moto(id="M001", data_registro="2026-05-28", quem_registrou="Emerson")
    build_workbook(motos=[moto], path=out)
    wb = load_workbook(out)
    ws = wb["Movimentações"]
    assert ws.max_row == 2
    assert ws["B2"].value == "M001"
    assert ws["D2"].value == "Aguardando diagnóstico"
    assert "Recebida" in (ws["F2"].value or "")
